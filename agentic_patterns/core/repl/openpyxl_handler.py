"""Handles serialization of openpyxl Workbook objects across subprocess boundaries."""

from pathlib import Path

from pydantic import BaseModel


class WorkbookReference(BaseModel):
    """Reference to a workbook saved in a temp file for IPC."""

    ref_type: str = "WorkbookReference"
    temp_path: Path
    var_name: str


def filter_openpyxl_from_namespace(
    namespace: dict, temp_dir: Path
) -> tuple[dict, set[str], list[str]]:
    """Filter openpyxl objects from namespace, serializing saveable workbooks.

    Returns:
        - dict of (var_name -> WorkbookReference dict) for saveable workbooks
        - set of keys that were handled (caller should skip these)
        - list of messages to display to user
    """
    result = {}
    handled_keys = set()
    messages = []

    for key, value in namespace.items():
        if is_saveable_workbook(value):
            wb_ref = save_workbook(value, key, temp_dir)
            result[key] = wb_ref.model_dump()
            handled_keys.add(key)
            continue

        if is_openpyxl_workbook(value):
            mode = "read-only" if getattr(value, "read_only", False) else "write-only"
            messages.append(
                f"Note: {mode} workbook '{key}' not persisted (cannot be serialized)"
            )
            handled_keys.add(key)
            continue

        if is_openpyxl_workbook_reference(value):
            wb = load_workbook_from_reference(value)
            wb_ref = save_workbook(wb, key, temp_dir)
            result[key] = wb_ref.model_dump()
            handled_keys.add(key)
            continue

        if is_openpyxl_worksheet(value):
            parent_wb = getattr(value, "parent", None)
            wb_var = None
            if parent_wb is not None:
                for k, v in namespace.items():
                    if v is parent_wb:
                        wb_var = k
                        break
            if wb_var:
                messages.append(
                    f"Note: worksheet '{key}' not persisted - re-access via: {key} = {wb_var}.active"
                )
            else:
                messages.append(
                    f"Note: worksheet '{key}' not persisted - re-access via workbook in next cell"
                )
            handled_keys.add(key)
            continue

        if is_openpyxl_object(value):
            handled_keys.add(key)
            continue

    return result, handled_keys, messages


def is_openpyxl_object(obj) -> bool:
    """Check if an object is any openpyxl type."""
    module_name = getattr(type(obj), "__module__", "")
    return module_name.startswith("openpyxl.")


def is_openpyxl_workbook(obj) -> bool:
    """Check if an object is an openpyxl Workbook."""
    obj_type = type(obj)
    if obj_type.__name__ != "Workbook":
        return False
    module_name = getattr(obj_type, "__module__", "")
    if not module_name.startswith("openpyxl."):
        return False
    return hasattr(obj, "active") and hasattr(obj, "save")


def is_openpyxl_workbook_reference(obj) -> bool:
    """Check if an object is a WorkbookReference (object or dict form)."""
    if isinstance(obj, WorkbookReference):
        return True
    if isinstance(obj, dict) and obj.get("ref_type") == "WorkbookReference":
        return True
    return False


def is_openpyxl_worksheet(obj) -> bool:
    """Check if an object is an openpyxl Worksheet."""
    obj_type = type(obj)
    if obj_type.__name__ != "Worksheet":
        return False
    module_name = getattr(obj_type, "__module__", "")
    if not module_name.startswith("openpyxl."):
        return False
    return hasattr(obj, "parent") and hasattr(obj, "cell")


def is_saveable_workbook(workbook) -> bool:
    """Check if a workbook can be saved."""
    if not is_openpyxl_workbook(workbook):
        return False
    if getattr(workbook, "read_only", False):
        return False
    if getattr(workbook, "write_only", False):
        return False
    return True


def load_workbook_from_reference(ref) -> "openpyxl.Workbook":  # type: ignore  # noqa: F821
    """Load a workbook from a reference (object or dict form)."""
    import openpyxl

    if isinstance(ref, WorkbookReference):
        return openpyxl.load_workbook(ref.temp_path)
    if isinstance(ref, dict) and "temp_path" in ref:
        return openpyxl.load_workbook(ref["temp_path"])
    raise ValueError(
        f"Expected WorkbookReference or dict with temp_path, got {type(ref)}"
    )


def restore_workbook_references(namespace: dict) -> None:
    """Restore workbook references in namespace by loading from temp files."""
    for key, value in list(namespace.items()):
        if is_openpyxl_workbook_reference(value):
            namespace[key] = load_workbook_from_reference(value)


def save_workbook(workbook, var_name: str, temp_dir: Path) -> WorkbookReference:
    """Save a workbook to a temp file and return a reference."""
    temp_path = temp_dir / f"{var_name}.xlsx"
    workbook.save(temp_path)
    return WorkbookReference(temp_path=temp_path, var_name=var_name)
