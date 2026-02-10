"""Spreadsheet file processor (.xlsx, .xls, .ods) with row/column truncation."""

import io
from collections.abc import Callable
from pathlib import Path

from agentic_patterns.core.context.config import ContextConfig, load_context_config
from agentic_patterns.core.context.models import (
    FileExtractionResult,
    FileType,
    TruncationInfo,
)
from agentic_patterns.core.context.processors.common import (
    check_and_apply_output_limit,
    create_error_result,
    create_file_metadata,
    format_section_header,
    truncate_string,
)


def _format_sheet_as_csv(
    rows: list[list], max_cell_length: int, max_line_length: int
) -> tuple[str, int]:
    """Format sheet rows as CSV with cell and line truncation."""
    output = io.StringIO()
    cells_truncated = 0

    for row in rows:
        truncated_cells = []
        for cell in row:
            cell_str = str(cell) if cell is not None else ""
            truncated_cell, was_truncated = truncate_string(cell_str, max_cell_length)
            truncated_cells.append(truncated_cell)
            if was_truncated:
                cells_truncated += 1

        line = ",".join(
            f'"{cell}"' if "," in cell or '"' in cell else cell
            for cell in truncated_cells
        )

        if len(line) > max_line_length:
            line = line[:max_line_length] + "..."

        output.write(line + "\n")

    return output.getvalue(), cells_truncated


def _process_excel_file(
    file_path: Path, config: ContextConfig, max_sheets: int
) -> tuple[str, TruncationInfo]:
    """Process Excel file (.xlsx) using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel file processing. Install with: uv add openpyxl"
        )

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0

    for i in range(sheets_to_process):
        sheet_name = sheet_names[i]
        sheet = workbook[sheet_name]

        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        if sheet.max_row is None or sheet.max_row == 0:
            output.write("(empty sheet)\n")
            continue

        total_rows = sheet.max_row - 1
        total_columns = sheet.max_column or 0
        columns_to_show = min(total_columns, config.max_columns)

        # Read header
        header = [
            str(sheet.cell(1, col).value or "") for col in range(1, columns_to_show + 1)
        ]

        # Select rows (head + tail pattern)
        if total_rows <= (config.rows_head + config.rows_tail):
            row_indices = list(range(2, sheet.max_row + 1))
        else:
            head_indices = list(range(2, config.rows_head + 2))
            tail_indices = (
                list(range(sheet.max_row - config.rows_tail + 1, sheet.max_row + 1))
                if config.rows_tail > 0
                else []
            )
            row_indices = head_indices + tail_indices

        # Build rows
        rows = [header]
        for row_idx in row_indices:
            row = [
                str(sheet.cell(row_idx, col).value or "")
                for col in range(1, columns_to_show + 1)
            ]
            rows.append(row)

        # Format as CSV
        csv_content, cells_truncated = _format_sheet_as_csv(
            rows, config.max_cell_length, config.max_line_length
        )
        total_cells_truncated += cells_truncated

        rows_shown = len(row_indices)
        output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")
        output.write(csv_content)

        if output.tell() > config.max_total_output:
            output.write("\n...\n")
            break

    workbook.close()

    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated,
        total_output_limit_reached=output.tell() > config.max_total_output,
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = (
            f"{sheets_to_process} sheets of {total_sheets} total"
        )

    return output.getvalue(), truncation_info


def _process_xls_file(
    file_path: Path, config: ContextConfig, max_sheets: int
) -> tuple[str, TruncationInfo]:
    """Process legacy Excel file (.xls) using xlrd."""
    try:
        import xlrd
    except ImportError:
        raise ImportError(
            "xlrd is required for .xls file processing. Install with: uv add xlrd"
        )

    workbook = xlrd.open_workbook(file_path)
    sheet_names = workbook.sheet_names()
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0

    for i in range(sheets_to_process):
        sheet_name = sheet_names[i]
        sheet = workbook.sheet_by_name(sheet_name)

        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        if sheet.nrows == 0:
            output.write("(empty sheet)\n")
            continue

        total_rows = sheet.nrows - 1
        total_columns = sheet.ncols
        columns_to_show = min(total_columns, config.max_columns)

        header = [str(sheet.cell_value(0, col)) for col in range(columns_to_show)]

        if total_rows <= (config.rows_head + config.rows_tail):
            row_indices = list(range(1, sheet.nrows))
        else:
            head_indices = list(range(1, config.rows_head + 1))
            tail_indices = (
                list(range(sheet.nrows - config.rows_tail, sheet.nrows))
                if config.rows_tail > 0
                else []
            )
            row_indices = head_indices + tail_indices

        rows = [header]
        for row_idx in row_indices:
            row = [
                str(sheet.cell_value(row_idx, col)) for col in range(columns_to_show)
            ]
            rows.append(row)

        csv_content, cells_truncated = _format_sheet_as_csv(
            rows, config.max_cell_length, config.max_line_length
        )
        total_cells_truncated += cells_truncated

        rows_shown = len(row_indices)
        output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")
        output.write(csv_content)

        if output.tell() > config.max_total_output:
            output.write("\n...\n")
            break

    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated,
        total_output_limit_reached=output.tell() > config.max_total_output,
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = (
            f"{sheets_to_process} sheets of {total_sheets} total"
        )

    return output.getvalue(), truncation_info


def process_spreadsheet(
    file_path: Path,
    max_sheets: int = 3,
    config: ContextConfig | None = None,
    tokenizer: Callable[[str], int] | None = None,
) -> FileExtractionResult:
    """Process spreadsheet files (.xlsx, .xls)."""
    if config is None:
        config = load_context_config()

    try:
        metadata = create_file_metadata(
            file_path, mime_type=f"spreadsheet/{file_path.suffix[1:]}"
        )

        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            content, truncation_info = _process_excel_file(
                file_path, config, max_sheets
            )
        elif suffix == ".xls":
            content, truncation_info = _process_xls_file(file_path, config, max_sheets)
        else:
            return FileExtractionResult(
                content=None,
                success=False,
                error_message=f"Unsupported spreadsheet format: {suffix}",
                file_type=FileType.SPREADSHEET,
                metadata=metadata,
            )

        if tokenizer:
            truncation_info.tokens_shown = tokenizer(content)

        content = check_and_apply_output_limit(
            content, config.max_total_output, truncation_info
        )

        return FileExtractionResult(
            content=content,
            success=True,
            was_extracted=True,
            file_type=FileType.SPREADSHEET,
            truncation_info=truncation_info,
            metadata=metadata,
        )

    except ImportError as e:
        return create_error_result(
            e, FileType.SPREADSHEET, file_path, "spreadsheet (missing dependencies)"
        )
    except Exception as e:
        return create_error_result(e, FileType.SPREADSHEET, file_path, "spreadsheet")
